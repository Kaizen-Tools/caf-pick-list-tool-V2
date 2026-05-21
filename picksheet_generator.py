"""Data validation, consolidation, and export helpers for CAF pick sheets."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import importlib.util
import re
from typing import BinaryIO

import pandas as pd

from location_parser import parse_location


REQUIRED_COLUMNS = ["Line", "Part", "Description", "Issue Qty.", "From Bin"]
OPTIONAL_OUTPUT_COLUMNS = [
    "Part Org.",
    "Condition",
    "Lot",
    "Expiration Date",
    "Serial Number",
    "Manufacturer",
    "Manufacturer Part Number",
]

OUTPUT_COLUMNS = [
    "Issue Reference",
    "Requisition Reference",
    "Line",
    "Part",
    "Description",
    "Issue Qty.",
    "From Bin",
    "Parsed Aisle",
    "Parsed Bay",
    "Parsed Vertical",
    "Parsed Lateral",
    "Pick Zone",
    "Source File",
]


@dataclass
class ValidationResult:
    filename: str
    missing_columns: list[str]
    row_count: int

    @property
    def is_valid(self) -> bool:
        return not self.missing_columns


def read_issue_export(file_obj: BinaryIO, filename: str) -> pd.DataFrame:
    """Read a GMAO STS Issue export without relying on column positions."""

    try:
        df = pd.read_excel(file_obj)
    except Exception as exc:  # pragma: no cover - Streamlit displays this text.
        raise ValueError(f"{filename}: could not read Excel file ({exc})") from exc

    df.columns = [str(col).strip() for col in df.columns]
    df = df.dropna(how="all").copy()
    return df


def validate_issue_export(df: pd.DataFrame, filename: str) -> ValidationResult:
    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    return ValidationResult(filename=filename, missing_columns=missing, row_count=len(df))


def consolidate_issue_exports(file_records: list[dict]) -> tuple[pd.DataFrame, list[ValidationResult]]:
    """Combine uploaded issue files and add user-entered traceability metadata."""

    frames: list[pd.DataFrame] = []
    validations: list[ValidationResult] = []

    for record in file_records:
        uploaded_file = record["file"]
        issue_reference = record.get("issue_reference", "")
        requisition_reference = record.get("requisition_reference", "")

        df = read_issue_export(uploaded_file, uploaded_file.name)
        validation = validate_issue_export(df, uploaded_file.name)
        validations.append(validation)
        if not validation.is_valid:
            continue

        df["Issue Reference"] = issue_reference
        df["Requisition Reference"] = requisition_reference
        df["Source File"] = uploaded_file.name
        frames.append(df)

    if not frames:
        return pd.DataFrame(), validations

    return pd.concat(frames, ignore_index=True), validations


def build_pick_sheets(consolidated: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Parse, classify, sort, and split consolidated issue lines."""

    if consolidated.empty:
        empty = pd.DataFrame(columns=OUTPUT_COLUMNS + ["Exception Reason"])
        return {"ground": empty, "height": empty, "exceptions": empty, "all": empty}

    df = consolidated.copy()
    parsed = df["From Bin"].apply(parse_location)

    df["Parsed Aisle"] = parsed.apply(lambda item: item.aisle)
    df["Parsed Bay"] = parsed.apply(lambda item: item.bay)
    df["Parsed Vertical"] = parsed.apply(lambda item: item.vertical)
    df["Parsed Lateral"] = parsed.apply(lambda item: item.lateral)
    df["Pick Zone"] = parsed.apply(lambda item: item.pick_zone)
    df["Exception Reason"] = parsed.apply(lambda item: item.exception_reason)
    df["_Route Aisle"] = parsed.apply(lambda item: item.route_key[0])
    df["_Route Bay"] = parsed.apply(lambda item: item.route_key[1])
    df["_Route Vertical"] = parsed.apply(lambda item: item.route_key[2])
    df["_Route Lateral Number"] = parsed.apply(lambda item: item.route_key[3])
    df["_Route Lateral"] = parsed.apply(lambda item: item.route_key[4])

    df = add_data_quality_exceptions(df)

    sort_cols = [
        "_Route Aisle",
        "_Route Bay",
        "_Route Vertical",
        "_Route Lateral Number",
        "_Route Lateral",
        "Part",
        "Line",
    ]
    df = df.sort_values(sort_cols, kind="stable", na_position="last")

    ground = clean_output(df[df["Pick Zone"] == "Ground-level"])
    height = clean_output(df[df["Pick Zone"] == "Height/FLT"])
    exceptions = clean_output(df[df["Pick Zone"] == "Exceptions"], include_exception=True)
    all_rows = clean_output(df, include_exception=True)

    return {"ground": ground, "height": height, "exceptions": exceptions, "all": all_rows}


def build_pick_packs(consolidated: pd.DataFrame, separate_issue_line_threshold: int) -> list[dict]:
    """Split uploaded issues into a consolidated pack plus separate large issue packs."""

    if consolidated.empty:
        return []

    threshold = _normalise_positive_int(separate_issue_line_threshold, default=20)
    issue_keys = _issue_keys(consolidated)
    issue_counts = issue_keys.value_counts(dropna=False)
    separate_issue_keys = set(issue_counts[issue_counts > threshold].index.tolist())

    packs = []
    consolidated_rows = consolidated.loc[~issue_keys.isin(separate_issue_keys)].copy()
    if not consolidated_rows.empty:
        packs.append(
            _build_pick_pack(
                name="Consolidated Pack",
                slug="consolidated-pack",
                pack_type="Consolidated",
                reason=f"Issues with {threshold} lines or fewer",
                rows=consolidated_rows,
            )
        )

    for issue_key in sorted(separate_issue_keys):
        issue_rows = consolidated.loc[issue_keys == issue_key].copy()
        issue_label = issue_key or "Not recorded"
        packs.append(
            _build_pick_pack(
                name=f"Separate Pack - {issue_label}",
                slug=f"separate-pack-{_slugify(issue_label)}",
                pack_type="Separate",
                reason=f"Issue exceeds {threshold} lines",
                rows=issue_rows,
            )
        )

    return packs


def _build_pick_pack(name: str, slug: str, pack_type: str, reason: str, rows: pd.DataFrame) -> dict:
    issue_references = sorted(value for value in _issue_keys(rows).drop_duplicates().tolist() if value)
    return {
        "name": name,
        "slug": slug,
        "pack_type": pack_type,
        "reason": reason,
        "line_count": len(rows),
        "issue_count": len(issue_references),
        "issue_references": issue_references,
        "pick_sheets": build_pick_sheets(rows),
    }


def _issue_keys(df: pd.DataFrame) -> pd.Series:
    return df["Requisition Reference"].fillna("").astype(str).str.strip()


def _normalise_positive_int(value: int, default: int) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return max(parsed, 1)


def _slugify(value: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", value.strip().lower()).strip("-")
    return slug or "not-recorded"


def add_data_quality_exceptions(df: pd.DataFrame) -> pd.DataFrame:
    """Flag rows that should not be forced into a pick sheet."""

    df = df.copy()
    required_data_checks = {
        "Part": "Missing part number",
        "Description": "Missing description",
        "Issue Qty.": "Missing issue quantity",
    }

    for column, reason in required_data_checks.items():
        missing = df[column].isna() | (df[column].astype(str).str.strip() == "")
        df.loc[missing, "Pick Zone"] = "Exceptions"
        df.loc[missing, "Exception Reason"] = df.loc[missing, "Exception Reason"].fillna(reason)

    duplicates = df.duplicated(
        subset=["Issue Reference", "Requisition Reference", "Source File", "Line", "Part", "From Bin"],
        keep=False,
    )
    duplicate_without_existing_reason = duplicates & df["Exception Reason"].isna()
    df.loc[duplicates, "Pick Zone"] = "Exceptions"
    df.loc[duplicate_without_existing_reason, "Exception Reason"] = "Potential duplicate line"

    return df


def clean_output(df: pd.DataFrame, include_exception: bool = False) -> pd.DataFrame:
    columns = OUTPUT_COLUMNS.copy()
    for column in OPTIONAL_OUTPUT_COLUMNS:
        if column in df.columns and column not in columns:
            columns.append(column)
    if include_exception:
        columns.append("Exception Reason")

    available = [col for col in columns if col in df.columns]
    return df.loc[:, available].reset_index(drop=True)


def dataframe_to_excel_bytes(sheets: dict[str, pd.DataFrame]) -> bytes:
    """Create a formatted Excel workbook in memory."""

    output = BytesIO()
    engine = "xlsxwriter" if importlib.util.find_spec("xlsxwriter") else "openpyxl"
    with pd.ExcelWriter(output, engine=engine) as writer:
        if engine == "xlsxwriter":
            _write_xlsxwriter_workbook(writer, sheets)
        else:
            _write_openpyxl_workbook(writer, sheets)

    return output.getvalue()


def _write_xlsxwriter_workbook(writer: pd.ExcelWriter, sheets: dict[str, pd.DataFrame]) -> None:
        workbook = writer.book
        header_format = workbook.add_format(
            {"bold": True, "font_color": "white", "bg_color": "#1F4E78", "border": 1}
        )
        date_format = workbook.add_format({"num_format": "yyyy-mm-dd"})

        for sheet_name, df in sheets.items():
            safe_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)
            worksheet = writer.sheets[safe_name]
            worksheet.freeze_panes(1, 0)
            worksheet.autofilter(0, 0, max(len(df), 1), max(len(df.columns) - 1, 0))

            for col_num, column in enumerate(df.columns):
                width = min(max(len(str(column)) + 2, 12), 36)
                if not df.empty:
                    sample_width = df[column].astype(str).str.len().quantile(0.95)
                    if pd.notna(sample_width):
                        width = min(max(width, int(sample_width) + 2), 48)
                fmt = date_format if "Date" in column else None
                worksheet.set_column(col_num, col_num, width, fmt)
                worksheet.write(0, col_num, column, header_format)


def _write_openpyxl_workbook(writer: pd.ExcelWriter, sheets: dict[str, pd.DataFrame]) -> None:
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    for sheet_name, df in sheets.items():
        safe_name = sheet_name[:31]
        df.to_excel(writer, sheet_name=safe_name, index=False)
        worksheet = writer.sheets[safe_name]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        for col_num, column in enumerate(df.columns, start=1):
            width = min(max(len(str(column)) + 2, 12), 36)
            if not df.empty:
                sample_width = df[column].astype(str).str.len().quantile(0.95)
                if pd.notna(sample_width):
                    width = min(max(width, int(sample_width) + 2), 48)
            worksheet.column_dimensions[get_column_letter(col_num)].width = width
